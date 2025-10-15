#!/usr/bin/env python3
"""
分析output.txt文件并生成Excel表格
功能：提取THD、SNR和幅频特性差异数据，按位深分类创建Excel表格
"""

import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import sys
import os

def parse_output_file(file_path):
    """
    解析output.txt文件，提取测试数据
    返回: dict {bit_depth: {src_rate: {dst_rate: {thd, snr, mag_diff}}}}
    """
    data = {}
    
    with open(file_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    # 匹配验证块的正则表达式
    # 格式: ==== 验证: filename (src_rate -> dst_rate, bits: XX) ====
    pattern = r'==== 验证: .+? \((\d+) -> (\d+), bits: (\d+)\) ====.*?thd_db: ([-\d.]+).*?snr_db: ([-\d.]+).*?magnitude_response_diff: ([-\d.]+)'
    
    matches = re.finditer(pattern, content, re.DOTALL)
    
    for match in matches:
        src_rate = int(match.group(1))
        dst_rate = int(match.group(2))
        bits = int(match.group(3))
        thd = float(match.group(4))
        snr = float(match.group(5))
        mag_diff = float(match.group(6))
        
        if bits not in data:
            data[bits] = {}
        if src_rate not in data[bits]:
            data[bits][src_rate] = {}
        
        data[bits][src_rate][dst_rate] = {
            'thd': thd,
            'snr': snr,
            'mag_diff': mag_diff
        }
    
    return data

def get_all_sample_rates(data):
    """获取所有出现的采样率"""
    all_rates = set()
    for bits in data.values():
        for src_rate in bits.keys():
            all_rates.add(src_rate)
            for dst_rate in bits[src_rate].keys():
                all_rates.add(dst_rate)
    return sorted(list(all_rates))

def create_dataframe(data, bits, metric, sample_rates):
    """
    创建DataFrame
    metric: 'thd', 'snr', 'mag_diff'
    """
    # 创建空DataFrame，使用字符串格式的采样率作为索引和列名
    df = pd.DataFrame(index=[str(r) for r in sample_rates], 
                     columns=[str(r) for r in sample_rates])
    
    # 填充数据
    if bits in data:
        for src_rate in sample_rates:
            if src_rate in data[bits]:
                for dst_rate in sample_rates:
                    if dst_rate in data[bits][src_rate]:
                        value = data[bits][src_rate][dst_rate][metric]
                        df.loc[str(src_rate), str(dst_rate)] = value
    
    return df

def format_excel_sheet(ws, title):
    """格式化Excel工作表"""
    # 设置标题行样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # 找到数据范围
    max_row = ws.max_row
    max_col = ws.max_column
    
    # 设置标题
    if max_col > 1:
        ws.merge_cells(f'A1:{get_column_letter(max_col)}1')
    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置列标题（第3行）
    if max_row >= 3:
        for col in range(2, max_col + 1):
            cell = ws.cell(row=3, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置行标题（第1列，从第4行开始）
    if max_row >= 4:
        for row in range(4, max_row + 1):
            cell = ws.cell(row=row, column=1)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 设置数据单元格格式
    for row in range(4, max_row + 1):
        for col in range(2, max_col + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                # 如果是数字，保留2位小数
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '0.00'
    
    # 自动调整列宽
    for col in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, max_row + 1):
            cell = ws.cell(row=row, column=col)
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 15)
        ws.column_dimensions[column_letter].width = adjusted_width

def create_excel(data, output_file):
    """创建Excel文件"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认工作表
    
    # 获取所有采样率
    all_sample_rates = get_all_sample_rates(data)
    
    # 为每个位深创建三个表格
    for bits in sorted(data.keys()):
        # THD表
        df_thd = create_dataframe(data, bits, 'thd', all_sample_rates)
        ws_thd = wb.create_sheet(title=f'{bits}bit_THD')
        ws_thd.append([f'{bits}bit THD (单位: dB)'])
        ws_thd.append([])  # 空行
        # 添加列标题行
        header_row = ['源采样率 →'] + [str(rate) for rate in all_sample_rates]
        ws_thd.append(header_row)
        # 添加数据行
        for src_rate in all_sample_rates:
            row_data = [str(src_rate)]
            for dst_rate in all_sample_rates:
                value = df_thd.loc[str(src_rate), str(dst_rate)]
                row_data.append(value if pd.notna(value) else '')
            ws_thd.append(row_data)
        format_excel_sheet(ws_thd, f'{bits}bit THD (单位: dB)')
        
        # SNR表
        df_snr = create_dataframe(data, bits, 'snr', all_sample_rates)
        ws_snr = wb.create_sheet(title=f'{bits}bit_SNR')
        ws_snr.append([f'{bits}bit SNR (单位: dB)'])
        ws_snr.append([])  # 空行
        # 添加列标题行
        header_row = ['源采样率 →'] + [str(rate) for rate in all_sample_rates]
        ws_snr.append(header_row)
        # 添加数据行
        for src_rate in all_sample_rates:
            row_data = [str(src_rate)]
            for dst_rate in all_sample_rates:
                value = df_snr.loc[str(src_rate), str(dst_rate)]
                row_data.append(value if pd.notna(value) else '')
            ws_snr.append(row_data)
        format_excel_sheet(ws_snr, f'{bits}bit SNR (单位: dB)')
        
        # 幅频特性差异表
        df_mag = create_dataframe(data, bits, 'mag_diff', all_sample_rates)
        ws_mag = wb.create_sheet(title=f'{bits}bit_幅频特性差异')
        ws_mag.append([f'{bits}bit 幅频特性差异 (单位: dB)'])
        ws_mag.append([])  # 空行
        # 添加列标题行
        header_row = ['源采样率 →'] + [str(rate) for rate in all_sample_rates]
        ws_mag.append(header_row)
        # 添加数据行
        for src_rate in all_sample_rates:
            row_data = [str(src_rate)]
            for dst_rate in all_sample_rates:
                value = df_mag.loc[str(src_rate), str(dst_rate)]
                row_data.append(value if pd.notna(value) else '')
            ws_mag.append(row_data)
        format_excel_sheet(ws_mag, f'{bits}bit 幅频特性差异 (单位: dB)')
    
    wb.save(output_file)
    print(f"Excel文件已保存: {output_file}")

def main():
    # 默认输入和输出文件路径
    input_file = 'output.txt'
    output_file = 'test_results.xlsx'
    
    # 如果提供了命令行参数
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
    if len(sys.argv) > 2:
        output_file = sys.argv[2]
    
    # 检查输入文件是否存在
    if not os.path.exists(input_file):
        print(f"错误: 找不到文件 {input_file}")
        sys.exit(1)
    
    print(f"正在解析文件: {input_file}")
    data = parse_output_file(input_file)
    
    if not data:
        print("错误: 未能从文件中提取到任何数据")
        sys.exit(1)
    
    print(f"找到 {len(data)} 个位深的数据:")
    for bits in sorted(data.keys()):
        print(f"  {bits}bit: {len(data[bits])} 个源采样率")
    
    print(f"正在创建Excel文件: {output_file}")
    create_excel(data, output_file)
    
    print("完成!")

if __name__ == '__main__':
    main()

